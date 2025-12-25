import csv
import re
from datetime import datetime, timedelta, time
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Shift Configuration
SHIFTS = {
    'General': {
        'short_name': 'GS',
        'begin_time': time(9, 0),
        'end_time': time(18, 0),
        'breaks': []
    },
    'Maintenance Shift': {
        'short_name': 'MS',
        'begin_time': time(7, 0),
        'end_time': time(16, 1),
        'breaks': []
    },
    'Shift': {
        'short_name': 'Shift',
        'begin_time': time(9, 0),
        'end_time': time(20, 30),
        'breaks': [
            (time(10, 45), time(11, 0)),   # Break 1: 15 mins
            (time(13, 0), time(13, 30)),   # Break 2: 30 mins
            (time(15, 45), time(16, 0)),   # Break 3: 15 mins
            (time(20, 30), time(21, 0))    # Break 4: 30 mins
        ]
    }
}

# Category/Employee Rules
CATEGORY_RULES = {
    'Cont Worked': {
        'grace_late_coming': 10,  # mins
        'grace_early_going': 10,  # mins
        'deduct_break_hours': True,
        'consider_late_going': True,
        'allow_midnight_shift': True,
        'punch_end_after': 450  # mins (7.5 hours) - threshold for next day detection
    }
}

class AttendanceProcessor:
    def __init__(self):
        self.employees = defaultdict(lambda: {
            'name': '',
            'company': '',
            'department': '',
            'day1_punches': [],
            'day2_punches': [],
            'day1_hours': 0,
            'day2_hours': 0,
            'total_hours': 0,
            'cross_day_shift': None  # Track shifts spanning midnight
        })

    def parse_time(self, time_str):
        """Convert time string (HH:MM) to datetime.time object"""
        try:
            return datetime.strptime(time_str.strip(), '%H:%M').time()
        except:
            return None

    def time_to_minutes(self, t):
        """Convert time object to minutes since midnight"""
        if t is None:
            return 0
        return t.hour * 60 + t.minute

    def minutes_to_time_str(self, minutes):
        """Convert minutes to HH:MM format"""
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours:02d}:{mins:02d}"

    def calculate_duration(self, entry_time, exit_time):
        """
        Calculate duration between entry and exit times.
        Handles day spanning (e.g., 23:00 to 01:00 next day) **only if allowed by config**.
        Returns minutes worked or 0 if cross-midnight shifts are not allowed.
        """
        entry_minutes = self.time_to_minutes(entry_time)
        exit_minutes = self.time_to_minutes(exit_time)

        # Check if cross-midnight shifts are allowed
        allow_midnight = CATEGORY_RULES.get('Cont Worked', {}).get('allow_midnight_shift', True)

        if exit_minutes >= entry_minutes:
            # Normal case: same day
            return exit_minutes - entry_minutes
        else:
            # Cross-midnight case
            if allow_midnight:
            # Count duration across midnight
                return (24 * 60 - entry_minutes) + exit_minutes
            else:
                # Cross-midnight shifts NOT allowed - ignore duration or handle as needed
                return 0

    def deduct_breaks(self, entry_time, exit_time, breaks):
        """
        Deduct break times from work duration.
        Breaks: list of tuples (break_start, break_end) in time objects
        Returns minutes to deduct
        """
        total_break_mins = 0
        entry_mins = self.time_to_minutes(entry_time)
        exit_mins = self.time_to_minutes(exit_time)

        # Handle day spanning for exit time
        if exit_mins < entry_mins:
            exit_mins += 24 * 60

        for break_start, break_end in breaks:
            break_start_mins = self.time_to_minutes(break_start)
            break_end_mins = self.time_to_minutes(break_end)

            # Check if break falls within work period
            if break_start_mins >= entry_mins and break_end_mins <= exit_mins:
                total_break_mins += (break_end_mins - break_start_mins)
            elif break_start_mins < exit_mins and break_end_mins > entry_mins:
                # Partial overlap
                overlap_start = max(break_start_mins, entry_mins)
                overlap_end = min(break_end_mins, exit_mins)
                if overlap_end > overlap_start:
                    total_break_mins += (overlap_end - overlap_start)

        return total_break_mins

    def apply_grace_time(self, entry_time, exit_time, shift_config, rules):
        """
        Apply grace time for late coming and early going.
        Returns adjusted work duration in minutes.
        """
        grace_late = rules.get('grace_late_coming', 0)
        grace_early = rules.get('grace_early_going', 0)

        shift_start = shift_config['begin_time']
        shift_end = shift_config['end_time']

        entry_mins = self.time_to_minutes(entry_time)
        exit_mins = self.time_to_minutes(exit_time)
        shift_start_mins = self.time_to_minutes(shift_start)
        shift_end_mins = self.time_to_minutes(shift_end)

        # Handle day spanning
        if exit_mins < entry_mins:
            exit_mins += 24 * 60
        if shift_end_mins < shift_start_mins:
            shift_end_mins += 24 * 60

        # Apply grace for late coming
        if entry_mins > shift_start_mins:
            late_mins = entry_mins - shift_start_mins
            if late_mins <= grace_late:
                entry_mins = shift_start_mins

        # Apply grace for early going
        if exit_mins < shift_end_mins:
            early_mins = shift_end_mins - exit_mins
            if early_mins <= grace_early:
                exit_mins = shift_end_mins

        return max(0, exit_mins - entry_mins)

    def parse_punch_records(self, punch_str):
        """Parse punch records string and return list of times"""
        if not punch_str or punch_str.strip() == '':
            return []

        # Split by comma and filter empty values
        punches = [p.strip() for p in punch_str.split(',') if p.strip()]
        times = []

        for punch in punches:
            time_obj = self.parse_time(punch)
            if time_obj:
                times.append(time_obj)

        return times

    def parse_breaks(self, breaks_str):
        """
        Parse breaks string and return list of (start_time, end_time) tuples.
        Format: 'HH:MM-HH:MM,HH:MM-HH:MM'
        Example: '10:30-11:00,13:00-13:30'
        Returns: [(time(10,30), time(11,0)), (time(13,0), time(13,30))]
        """
        if not breaks_str or breaks_str.strip() == '':
            return []

        breaks = []
        break_periods = [b.strip() for b in breaks_str.split(',') if b.strip()]

        for period in break_periods:
            if '-' not in period:
                raise ValueError(f"Invalid break format: '{period}'. Expected 'HH:MM-HH:MM'")

            parts = period.split('-')
            if len(parts) != 2:
                raise ValueError(f"Invalid break format: '{period}'. Expected 'HH:MM-HH:MM'")

            start_str = parts[0].strip()
            end_str = parts[1].strip()

            start_time = self.parse_time(start_str)
            end_time = self.parse_time(end_str)

            if not start_time or not end_time:
                raise ValueError(f"Invalid time format in break: '{period}'")

            breaks.append((start_time, end_time))

        return breaks

    def minutes_to_decimal_hours(self, minutes):
        """Convert minutes to decimal hours"""
        return round(minutes / 60, 2)

    def calculate_working_hours(self, punches, day='day2'):
        """
        Calculate total working hours from punch records.
        Assumes punches are in order: in, out, in, out, ...
        Applies shift rules including breaks and grace time.
        For day 1: Uses 09:00 as reference point - only counts time from 09:00 onwards.
        For day 2: Uses actual punch times.
        Returns total minutes worked.
        """
        if len(punches) < 2:
            return 0

        # Remove all duplicate punches (keep first occurrence only)
        seen = set()
        deduplicated = []
        for punch in punches:
            if punch not in seen:
                seen.add(punch)
                deduplicated.append(punch)

        if len(deduplicated) < 2:
            return 0

        # Use default shift configuration
        shift_config = SHIFTS['Shift']
        rules = CATEGORY_RULES['Cont Worked']

        # Reference time for Day 1: 09:00
        reference_time = time(9, 0)
        reference_mins = self.time_to_minutes(reference_time)
        use_reference = (day == 'day1')

        total_minutes = 0
        # Process pairs of punches (entry, exit)
        for i in range(0, len(deduplicated) - 1, 2):
            entry = deduplicated[i]
            exit_time = deduplicated[i + 1]

            # For Day 1, adjust entry time to be at least 09:00
            if use_reference:
                entry_mins = self.time_to_minutes(entry)
                if entry_mins < reference_mins:
                    # Clock in was before 09:00, adjust to 09:00
                    entry = reference_time

            # Calculate raw duration
            duration = self.calculate_duration(entry, exit_time)

            # Deduct breaks if configured
            if rules.get('deduct_break_hours', True):
                break_mins = self.deduct_breaks(entry, exit_time, shift_config['breaks'])
                duration -= break_mins

            # Apply grace time
            grace_adjusted = self.apply_grace_time(entry, exit_time, shift_config, rules)
            if grace_adjusted > 0:
                duration = grace_adjusted
                # Re-deduct breaks after grace adjustment
                if rules.get('deduct_break_hours', True):
                    break_mins = self.deduct_breaks(entry, exit_time, shift_config['breaks'])
                    duration -= break_mins

            if duration > 0:
                total_minutes += duration

        return max(0, total_minutes)

    def calculate_working_hours_total_span(self, punches, day='day2'):
        """
        Calculate total working hours from first punch to last punch.
        Considers all punches as a continuous span.
        Deducts breaks that fall within the total span.
        For Day 1: Filters out all punches before 09:00 AM (start reference point)
        Returns total minutes worked.
        """
        if len(punches) < 2:
            return 0

        # Remove all duplicate punches (keep first occurrence only)
        seen = set()
        deduplicated = []
        for punch in punches:
            if punch not in seen:
                seen.add(punch)
                deduplicated.append(punch)

        # Reference time for Day 1: 09:00
        reference_time = time(9, 0)
        reference_mins = self.time_to_minutes(reference_time)

        # For Day 1, filter out all punches before 09:00
        if day == 'day1':
            filtered = []
            for punch in deduplicated:
                punch_mins = self.time_to_minutes(punch)
                if punch_mins >= reference_mins:
                    filtered.append(punch)
            deduplicated = filtered

        if len(deduplicated) < 2:
            return 0

        # Get first and last punch times
        first_punch = deduplicated[0]
        last_punch = deduplicated[-1]

        # Use default shift configuration
        shift_config = SHIFTS['Shift']
        rules = CATEGORY_RULES['Cont Worked']

        # Calculate total duration from first to last punch
        total_duration = self.calculate_duration(first_punch, last_punch)

        # Deduct breaks if configured
        if rules.get('deduct_break_hours', True):
            break_mins = self.deduct_breaks(first_punch, last_punch, shift_config['breaks'])
            total_duration -= break_mins

        return max(0, total_duration)

    def analyze_punch_pairs(self, punches, custom_breaks=None):
        """
        Analyze punch pairs and identify unpaired punches with detailed breakdown.
        custom_breaks: List of tuples like [(time(10,45), time(11,0)), (time(15,45), time(16,0))]
        If None, uses SHIFTS['Shift']['breaks']
        Returns a dictionary with analysis results.
        """
        if not punches:
            return {'total_punches': 0, 'pairs': [], 'unpaired': None, 'total_minutes': 0}

        # Remove all duplicate punches (keep first occurrence only)
        seen = set()
        deduplicated = []
        for punch in punches:
            if punch not in seen:
                seen.add(punch)
                deduplicated.append(punch)

        shift_config = SHIFTS['Shift']

        # Use custom breaks if provided, otherwise use shift config breaks
        breaks = custom_breaks if custom_breaks is not None else shift_config['breaks']

        pairs = []
        total_minutes = 0

        # Process complete pairs
        for i in range(0, len(deduplicated) - 1, 2):
            entry = deduplicated[i]
            exit_time = deduplicated[i + 1]

            duration = self.calculate_duration(entry, exit_time)
            break_mins = self.deduct_breaks(entry, exit_time, breaks)
            final_duration = duration - break_mins

            pairs.append({
                'pair_num': (i // 2) + 1,
                'entry': entry,
                'exit': exit_time,
                'raw_duration_mins': duration,
                'breaks_deducted_mins': break_mins,
                'final_duration_mins': final_duration,
                'final_duration_hrs': round(final_duration / 60, 2)
            })

            if final_duration > 0:
                total_minutes += final_duration

        # Check for unpaired punch
        unpaired = None
        if len(deduplicated) % 2 == 1:
            unpaired = {
                'punch': deduplicated[-1],
                'position': len(deduplicated),
                'likely_type': 'Clock IN (entry)' if len(pairs) % 2 == 0 else 'Clock OUT (exit)',
                'status': 'Missing exit/entry punch',
                'unaccounted_mins': 'Unknown - cannot calculate without exit/entry'
            }

        return {
            'total_punches': len(deduplicated),
            'complete_pairs': len(pairs),
            'pairs': pairs,
            'unpaired': unpaired,
            'total_minutes': total_minutes,
            'total_hours': round(total_minutes / 60, 2),
            'breaks_config': breaks
        }

    def print_punch_analysis(self, punches, label="", custom_breaks=None):
        """
        Print detailed analysis of punches in a formatted table.
        custom_breaks: List of tuples like [(time(10,45), time(11,0)), (time(15,45), time(16,0))]
        """
        analysis = self.analyze_punch_pairs(punches, custom_breaks)

        print("\n" + "=" * 100)
        if label:
            print(f"PUNCH ANALYSIS: {label}")
        else:
            print("PUNCH ANALYSIS")
        print("=" * 100)
        print(f"\nTotal Punches: {analysis['total_punches']} | Complete Pairs: {analysis['complete_pairs']}")

        # Display breaks configuration
        if analysis['breaks_config']:
            breaks_str = ", ".join([f"{b[0]}-{b[1]}" for b in analysis['breaks_config']])
            print(f"Breaks Configuration: {breaks_str}")
        else:
            print("Breaks Configuration: No breaks")

        print("-" * 100)
        print(f"{'Pair':<6} {'IN Time':<12} {'OUT Time':<12} {'Raw':<8} {'Breaks':<8} {'Final':<8} {'Hours':<8}")
        print("-" * 100)

        for pair in analysis['pairs']:
            print(f"{pair['pair_num']:<6} {str(pair['entry']):<12} {str(pair['exit']):<12} "
                  f"{pair['raw_duration_mins']:<8} {pair['breaks_deducted_mins']:<8} "
                  f"{pair['final_duration_mins']:<8} {pair['final_duration_hrs']:<8}")

        if analysis['unpaired']:
            unpaired = analysis['unpaired']
            print("-" * 100)
            print(f"\nUNPAIRED PUNCH:")
            print(f"  Time: {unpaired['punch']}")
            print(f"  Type: {unpaired['likely_type']}")
            print(f"  Status: {unpaired['status']}")
            print(f"  Unaccounted: {unpaired['unaccounted_mins']}")

        print("-" * 100)
        print(f"\nTOTAL WORKING TIME: {analysis['total_minutes']} minutes = {analysis['total_hours']} hours")
        print("=" * 100)

    def detect_cross_midnight_shift(self, day1_punches, day2_punches):
        """
        Detect if there's a shift spanning from late evening Day 1 to early morning Day 2 (before 6:00 AM).

        Logic:
        - Shift time starts at 9:00 AM
        - Any punches from midnight (00:00) up to 6:00 AM on Day 2 belong to the Day 1 shift
        - These early morning punches should be moved to Day 1
        - Punches from 6:00 AM onwards on Day 2 are considered Day 2 punches
        - Returns tuple of (adjusted_day1_punches, adjusted_day2_punches, cross_shift_info)
        """
        if not day1_punches or not day2_punches:
            return day1_punches, day2_punches, None

        midnight_boundary = time(7, 15)  # 6:45 AM boundary
        boundary_mins = self.time_to_minutes(midnight_boundary)

        # Check if Day 1 has punches
        if not day1_punches:
            return day1_punches, day2_punches, None

        # Check if Day 2 starts with punches before 6:00 AM (midnight cross-over)
        day2_early_punches = []
        day2_remaining_punches = []

        for punch in day2_punches:
            punch_mins = self.time_to_minutes(punch)
            if punch_mins < boundary_mins:  # Before 6:00 AM
                day2_early_punches.append(punch)
            else:
                day2_remaining_punches.append(punch)

        if day2_early_punches:
            # Early morning punches exist (before 6:00 AM) - they belong to the Day 1 shift
            # Move them to Day 1
            adjusted_day1 = day1_punches + day2_early_punches
            adjusted_day2 = day2_remaining_punches

            return adjusted_day1, adjusted_day2, {
                'early_morning_punches': day2_early_punches,
                'is_cross_midnight': True,
                'shifted_to_day1': True
            }

        return day1_punches, day2_punches, None

    def calculate_cross_midnight_hours(self, entry_time, exit_time):
        """
        Calculate hours for a shift that spans midnight.
        entry_time: time on Day 1
        exit_time: time on Day 2 (early morning)
        """
        shift_config = SHIFTS['Shift']
        rules = CATEGORY_RULES['Cont Worked']

        # Calculate raw duration from Day 1 evening to Day 2 morning
        duration = self.calculate_duration(entry_time, exit_time)

        # For cross-midnight shifts, deduct breaks if they fall within the window
        if rules.get('deduct_break_hours', True):
            break_mins = self.deduct_breaks(entry_time, exit_time, shift_config['breaks'])
            duration -= break_mins

        return max(0, duration)

    def read_attendance_file(self, filepath, day_key):
        """Read attendance file and populate employee data"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, quotechar='"')
                # Skip header
                next(reader)

                for row in reader:
                    # Each row is a single string with quotes, need to parse it manually
                    if len(row) == 1:
                        # Parse the quoted row
                        row_str = row[0]
                        # Extract fields between quotes
                        fields = []
                        in_quotes = False
                        current_field = ''
                        i = 0
                        while i < len(row_str):
                            if row_str[i] == '"':
                                if in_quotes and i + 1 < len(row_str) and row_str[i + 1] == '"':
                                    # Escaped quote
                                    current_field += '"'
                                    i += 2
                                else:
                                    # Toggle quote state
                                    in_quotes = not in_quotes
                                    i += 1
                            elif row_str[i] == ',' and not in_quotes:
                                # Field separator
                                fields.append(current_field)
                                current_field = ''
                                i += 1
                            else:
                                current_field += row_str[i]
                                i += 1
                        if current_field or in_quotes:
                            fields.append(current_field)

                        row = fields

                    if len(row) < 8:
                        continue

                    emp_code = row[0].strip()
                    emp_name = row[1].strip()
                    company = row[2].strip()
                    department = row[3].strip()
                    punch_str = row[6].strip()
                    status = row[7].strip()

                    if not emp_code or status == 'Not Present':
                        continue

                    # Store employee info
                    if day_key == 'day1':
                        self.employees[emp_code]['name'] = emp_name
                        self.employees[emp_code]['company'] = company
                        self.employees[emp_code]['department'] = department

                    # Parse punches and calculate hours
                    punches = self.parse_punch_records(punch_str)
                    hours = self.calculate_working_hours_total_span(punches, day_key)

                    if day_key == 'day1':
                        self.employees[emp_code]['day1_punches'] = punches
                        self.employees[emp_code]['day1_hours'] = hours
                    else:
                        self.employees[emp_code]['day2_punches'] = punches
                        self.employees[emp_code]['day2_hours'] = hours

        except Exception as e:
            print(f"Error reading {filepath}: {e}")
            import traceback
            traceback.print_exc()

    def generate_excel(self, output_filepath):
        """Generate Excel file with attendance summary"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = [
            "Employee Code",
            "Employee Name",
            "Company",
            "Department",
            "Day 1 Punches",
            "Day 1 Hours",
            "Day 2 Punches",
            "Day 2 Hours",
            "Total Hours"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Set column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12

        # Write data
        row = 2
        for emp_code in sorted(self.employees.keys()):
            emp = self.employees[emp_code]

            # Skip employees with no hours
            if emp['day1_hours'] == 0 and emp['day2_hours'] == 0:
                continue

            day1_hours_decimal = self.minutes_to_decimal_hours(emp['day1_hours'])
            day2_hours_decimal = self.minutes_to_decimal_hours(emp['day2_hours'])
            total_hours = day1_hours_decimal  # Total duration = day1 hours only

            # Format punch times
            day1_punches_str = ', '.join([t.strftime('%H:%M') for t in emp['day1_punches']]) if emp['day1_punches'] else ''
            day2_punches_str = ', '.join([t.strftime('%H:%M') for t in emp['day2_punches']]) if emp['day2_punches'] else ''

            # Write row
            ws.cell(row=row, column=1).value = emp_code
            ws.cell(row=row, column=2).value = emp['name']
            ws.cell(row=row, column=3).value = emp['company']
            ws.cell(row=row, column=4).value = emp['department']
            ws.cell(row=row, column=5).value = day1_punches_str
            ws.cell(row=row, column=6).value = day1_hours_decimal
            ws.cell(row=row, column=7).value = day2_punches_str
            ws.cell(row=row, column=8).value = day2_hours_decimal
            ws.cell(row=row, column=9).value = total_hours

            # Apply borders and alignment
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in [1, 3, 4, 5, 7]:  # Text columns
                    cell.alignment = data_alignment
                else:  # Number columns
                    cell.alignment = center_alignment
                    if col in [6, 8, 9]:  # Hour columns
                        cell.number_format = '0.00'

            row += 1

        # Freeze first row
        ws.freeze_panes = 'A2'

        # Save workbook
        wb.save(output_filepath)
        print(f"Excel file generated: {output_filepath}")

    def process(self, day1_file, day2_file, output_file):
        """Main processing method"""
        print("Reading Day 1 attendance...")
        self.read_attendance_file(day1_file, 'day1')

        print("Reading Day 2 attendance...")
        self.read_attendance_file(day2_file, 'day2')

        # Define the 9 AM reference threshold
        reference_time = time(9, 0)
        reference_mins = self.time_to_minutes(reference_time)
        early_boundary = time(7, 30)  # 7:30 AM boundary
        early_boundary_mins = self.time_to_minutes(early_boundary)

        # Detect and handle cross-midnight shifts
        print("Detecting cross-midnight shifts...")
        midnight_boundary = time(7, 15)  # 6:45 AM boundary
        boundary_mins = self.time_to_minutes(midnight_boundary)

        for emp_code in self.employees.keys():
            emp = self.employees[emp_code]
            day1_punches, day2_punches, cross_info = self.detect_cross_midnight_shift(
                emp['day1_punches'], emp['day2_punches']
            )

            # Update punches with adjusted versions
            emp['day1_punches'] = day1_punches
            emp['day2_punches'] = day2_punches

            # Filter and adjust Day 1 punches: punches between 7:30 AM to 8:59 AM are treated as 9:00 AM
            day1_filtered = []
            for punch in day1_punches:
                punch_mins = self.time_to_minutes(punch)
                if punch_mins >= early_boundary_mins:
                    # Punch is after 7:30 AM
                    if punch_mins < reference_mins:
                        # Punch is between 7:30 AM and 8:59 AM - treat as 9:00 AM
                        day1_filtered.append(reference_time)
                    else:
                        # Punch is at or after 9:00 AM - keep as is
                        day1_filtered.append(punch)

            # Filter day 2 punches up to 6:45 AM (early morning punches)
            day2_early_punches = []
            day2_remaining_punches = []
            for punch in day2_punches:
                punch_mins = self.time_to_minutes(punch)
                if punch_mins <= boundary_mins:
                    day2_early_punches.append(punch)
                else:
                    day2_remaining_punches.append(punch)

            # Combine Day 1 punches (from 7:30 AM onwards, with 7:30-8:59 treated as 9 AM) with Day 2 early morning punches (up to 6:45 AM)
            combined_punches = day1_filtered + day2_early_punches

            if len(combined_punches) >= 2:
                # Calculate total hours from Day 1 9:00 AM through Day 2 6:45 AM
                combined_hours = self.calculate_working_hours_total_span(combined_punches, 'day1')
                emp['day1_hours'] = combined_hours
                #emp['day2_hours'] = 0
            else:
                #emp['day1_hours'] = self.calculate_working_hours_total_span(day1_filtered, 'day1')
                emp['day2_hours'] = self.calculate_working_hours_total_span(day2_remaining_punches, 'day2')

            if cross_info:
                emp['cross_day_shift'] = cross_info

        print(f"Total employees processed: {len(self.employees)}")

        print("Generating Excel file...")
        self.generate_excel(output_file)

        print("Done!")

def main():
    import sys

    processor = AttendanceProcessor()

    # Check if running in analysis mode
    if len(sys.argv) > 1 and sys.argv[1] == '--analyze-punches':
        if len(sys.argv) < 3:
            print("Usage: python attendance_processor.py --analyze-punches '<punch_string>' [--breaks '<break_string>']")
            print("Example: python attendance_processor.py --analyze-punches '09:00,09:00,10:45,11:00,15:45,16:00,20:02,20:36'")
            print("With custom breaks: python attendance_processor.py --analyze-punches '08:50,00:01' --breaks '10:30-11:00,13:00-13:30'")
            print("For no breaks: python attendance_processor.py --analyze-punches '08:50,00:01' --breaks ''")
            sys.exit(1)

        punch_str = sys.argv[2]

        # Parse breaks parameter if provided
        custom_breaks = None
        if len(sys.argv) > 3 and sys.argv[3] == '--breaks':
            if len(sys.argv) > 4:
                breaks_str = sys.argv[4]
                if breaks_str.strip():  # If not empty string
                    try:
                        custom_breaks = processor.parse_breaks(breaks_str)
                    except Exception as e:
                        print(f"Error parsing breaks: {e}")
                        print("Break format should be: 'HH:MM-HH:MM,HH:MM-HH:MM'")
                        sys.exit(1)
                else:
                    custom_breaks = []  # Empty list for no breaks

        punches = processor.parse_punch_records(punch_str)
        processor.print_punch_analysis(punches, f"Custom Punches: {punch_str}", custom_breaks)
    else:
        # Normal mode: Process attendance files
        processor.process(
            r'c:\contract\day1.txt',
            r'c:\contract\day2.txt',
            r'c:\contract\Desiredoutput_final_new.xlsx'
        )

if __name__ == '__main__':
    main()
