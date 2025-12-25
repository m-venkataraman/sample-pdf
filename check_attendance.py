import csv
import re
from datetime import datetime, timedelta, time
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Shift Configuration
SHIFTS = {
    'General': {
        'short_name': 'GS',
        'begin_time': time(9, 0),
        'end_time': time(18, 0),
        'breaks': []
    },
    'Maintenance Shift': {
        'short_name': 'MS',
        'begin_time': time(7, 0),
        'end_time': time(16, 1),
        'breaks': []
    },
    'Shift': {
        'short_name': 'Shift',
        'begin_time': time(9, 0),
        'end_time': time(20, 30),
        'breaks': [
            (time(10, 45), time(11, 0)),   # Break 1: 15 mins
            (time(13, 0), time(13, 30)),   # Break 2: 30 mins
            (time(15, 45), time(16, 0)),   # Break 3: 15 mins
            (time(18, 0), time(18, 15)),   # Break 4: 15 mins
            (time(20, 30), time(21, 0))    # Break 5: 30 mins
        ]
    }
}

# Category/Employee Rules
CATEGORY_RULES = {
    'Cont Worked': {
        'grace_late_coming': 10,  # mins
        'grace_early_going': 10,  # mins
        'deduct_break_hours': True,
        'consider_late_going': True,
        'allow_midnight_shift': True,
        'punch_end_after': 450  # mins (7.5 hours) - threshold for next day detection
    }
}

class AttendanceProcessor:
    def __init__(self):
        self.employees = defaultdict(lambda: {
            'name': '',
            'company': '',
            'department': '',
            'day1_punches': [],
            'day2_punches': [],
            'day1_hours': 0,
            'day2_hours': 0,
            'total_hours': 0,
            'cross_day_shift': None
        })

    # ------------------------ Time Utilities ------------------------
    def parse_time(self, time_str):
        try:
            return datetime.strptime(time_str.strip(), '%H:%M').time()
        except:
            return None

    def time_to_minutes(self, t):
        if t is None:
            return 0
        return t.hour * 60 + t.minute

    def minutes_to_time_str(self, minutes):
        hours = minutes // 60
        mins = minutes % 60
        return f"{hours:02d}:{mins:02d}"

    def minutes_to_decimal_hours(self, minutes):
        return round(minutes / 60, 2)

    # ------------------------ Duration & Break Calculations ------------------------
    def calculate_duration(self, entry_time, exit_time):
        entry_minutes = self.time_to_minutes(entry_time)
        exit_minutes = self.time_to_minutes(exit_time)

        allow_midnight = CATEGORY_RULES.get('Cont Worked', {}).get('allow_midnight_shift', True)

        if exit_minutes >= entry_minutes:
            return exit_minutes - entry_minutes
        else:
            if allow_midnight:
                return (24 * 60 - entry_minutes) + exit_minutes
            else:
                return 0

    def deduct_breaks(self, entry_time, exit_time, breaks):
        total_break_mins = 0
        entry_mins = self.time_to_minutes(entry_time)
        exit_mins = self.time_to_minutes(exit_time)

        if exit_mins < entry_mins:
            exit_mins += 24 * 60

        for break_start, break_end in breaks:
            break_start_mins = self.time_to_minutes(break_start)
            break_end_mins = self.time_to_minutes(break_end)

            if break_start_mins >= entry_mins and break_end_mins <= exit_mins:
                total_break_mins += (break_end_mins - break_start_mins)
            elif break_start_mins < exit_mins and break_end_mins > entry_mins:
                overlap_start = max(break_start_mins, entry_mins)
                overlap_end = min(break_end_mins, exit_mins)
                if overlap_end > overlap_start:
                    total_break_mins += (overlap_end - overlap_start)

        return total_break_mins

    def apply_grace_time(self, entry_time, exit_time, shift_config, rules):
        grace_late = rules.get('grace_late_coming', 0)
        grace_early = rules.get('grace_early_going', 0)
        shift_start = shift_config['begin_time']
        shift_end = shift_config['end_time']

        entry_mins = self.time_to_minutes(entry_time)
        exit_mins = self.time_to_minutes(exit_time)
        shift_start_mins = self.time_to_minutes(shift_start)
        shift_end_mins = self.time_to_minutes(shift_end)

        if exit_mins < entry_mins:
            exit_mins += 24 * 60
        if shift_end_mins < shift_start_mins:
            shift_end_mins += 24 * 60

        if entry_mins > shift_start_mins:
            late_mins = entry_mins - shift_start_mins
            if late_mins <= grace_late:
                entry_mins = shift_start_mins

        if exit_mins < shift_end_mins:
            early_mins = shift_end_mins - exit_mins
            if early_mins <= grace_early:
                exit_mins = shift_end_mins

        return max(0, exit_mins - entry_mins)

    # ------------------------ Punch Parsing ------------------------
    def parse_punch_records(self, punch_str):
        if not punch_str or punch_str.strip() == '':
            return []

        punches = [p.strip() for p in punch_str.split(',') if p.strip()]
        times = []

        for punch in punches:
            time_obj = self.parse_time(punch)
            if time_obj:
                times.append(time_obj)

        return times

    def parse_breaks(self, breaks_str):
        if not breaks_str or breaks_str.strip() == '':
            return []

        breaks = []
        break_periods = [b.strip() for b in breaks_str.split(',') if b.strip()]

        for period in break_periods:
            if '-' not in period:
                raise ValueError(f"Invalid break format: '{period}'")
            parts = period.split('-')
            if len(parts) != 2:
                raise ValueError(f"Invalid break format: '{period}'")

            start_time = self.parse_time(parts[0].strip())
            end_time = self.parse_time(parts[1].strip())
            if not start_time or not end_time:
                raise ValueError(f"Invalid time format in break: '{period}'")

            breaks.append((start_time, end_time))

        return breaks

    # ------------------------ Hours Calculations ------------------------
    def calculate_working_hours_total_span(self, punches, day='day2'):
        if len(punches) < 2:
            return 0

        seen = set()
        deduplicated = []
        for punch in punches:
            if punch not in seen:
                seen.add(punch)
                deduplicated.append(punch)

        if len(deduplicated) < 2:
            return 0

        first_punch = deduplicated[0]
        last_punch = deduplicated[-1]

        shift_config = SHIFTS['Shift']
        rules = CATEGORY_RULES['Cont Worked']

        # ------------------- Day 1 rule: first punch before 9 AM -------------------
        if day == 'day1':
            nine_am = time(9, 0)
            if first_punch < nine_am:
                first_punch = nine_am

        total_duration = self.calculate_duration(first_punch, last_punch)

        if rules.get('deduct_break_hours', True):
            total_duration -= self.deduct_breaks(first_punch, last_punch, shift_config['breaks'])

        return max(0, total_duration)

    # ------------------------ Cross-Midnight Detection ------------------------
    def detect_cross_midnight_shift(self, day1_punches, day2_punches):
        if not day1_punches or not day2_punches:
            return day1_punches, day2_punches, None

        midnight_boundary = time(7, 15)
        boundary_mins = self.time_to_minutes(midnight_boundary)

        day2_early_punches = []
        day2_remaining_punches = []

        for punch in day2_punches:
            punch_mins = self.time_to_minutes(punch)
            if punch_mins < boundary_mins:
                day2_early_punches.append(punch)
            else:
                day2_remaining_punches.append(punch)

        if day2_early_punches:
            adjusted_day1 = day1_punches + day2_early_punches
            adjusted_day2 = day2_remaining_punches
            return adjusted_day1, adjusted_day2, {
                'early_morning_punches': day2_early_punches,
                'is_cross_midnight': True,
                'shifted_to_day1': True
            }

        return day1_punches, day2_punches, None

    # ------------------------ File Reading ------------------------
    def read_attendance_file(self, filepath, day_key):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, quotechar='"')
                next(reader)
                for row in reader:
                    if len(row) == 1:
                        row_str = row[0]
                        fields = []
                        in_quotes = False
                        current_field = ''
                        i = 0
                        while i < len(row_str):
                            if row_str[i] == '"':
                                if in_quotes and i + 1 < len(row_str) and row_str[i + 1] == '"':
                                    current_field += '"'
                                    i += 2
                                else:
                                    in_quotes = not in_quotes
                                    i += 1
                            elif row_str[i] == ',' and not in_quotes:
                                fields.append(current_field)
                                current_field = ''
                                i += 1
                            else:
                                current_field += row_str[i]
                                i += 1
                        if current_field or in_quotes:
                            fields.append(current_field)
                        row = fields

                    if len(row) < 8:
                        continue

                    emp_code = row[0].strip()
                    emp_name = row[1].strip()
                    company = row[2].strip()
                    department = row[3].strip()
                    punch_str = row[6].strip()
                    status = row[7].strip()

                    if not emp_code or status == 'Not Present':
                        continue

                    if day_key == 'day1':
                        self.employees[emp_code]['name'] = emp_name
                        self.employees[emp_code]['company'] = company
                        self.employees[emp_code]['department'] = department

                    punches = self.parse_punch_records(punch_str)
                    hours = self.calculate_working_hours_total_span(punches, day_key)

                    if day_key == 'day1':
                        self.employees[emp_code]['day1_punches'] = punches
                        self.employees[emp_code]['day1_hours'] = hours
                    else:
                        self.employees[emp_code]['day2_punches'] = punches
                        self.employees[emp_code]['day2_hours'] = hours
        except Exception as e:
            print(f"Error reading {filepath}: {e}")
            import traceback
            traceback.print_exc()

    # ------------------------ Excel Generation ------------------------
    def generate_excel(self, output_filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            "Employee Code", "Employee Name", "Company", "Department",
            "Day 1 Punches", "Day 1 Hours", "Day 2 Punches", "Day 2 Hours", "Total Hours"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12

        row = 2
        for emp_code in sorted(self.employees.keys()):
            emp = self.employees[emp_code]
            if emp['day1_hours'] == 0 and emp['day2_hours'] == 0:
                continue

            day1_hours_decimal = self.minutes_to_decimal_hours(emp['day1_hours'])
            day2_hours_decimal = self.minutes_to_decimal_hours(emp['day2_hours'])
            total_hours = day1_hours_decimal

            day1_punches_str = ', '.join([t.strftime('%H:%M') for t in emp['day1_punches']]) if emp['day1_punches'] else ''
            day2_punches_str = ', '.join([t.strftime('%H:%M') for t in emp['day2_punches']]) if emp['day2_punches'] else ''

            ws.cell(row=row, column=1).value = emp_code
            ws.cell(row=row, column=2).value = emp['name']
            ws.cell(row=row, column=3).value = emp['company']
            ws.cell(row=row, column=4).value = emp['department']
            ws.cell(row=row, column=5).value = day1_punches_str
            ws.cell(row=row, column=6).value = day1_hours_decimal
            ws.cell(row=row, column=7).value = day2_punches_str
            ws.cell(row=row, column=8).value = day2_hours_decimal
            ws.cell(row=row, column=9).value = total_hours

            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in [1, 3, 4, 5, 7]:
                    cell.alignment = data_alignment
                else:
                    cell.alignment = center_alignment
                    if col in [6, 8, 9]:
                        cell.number_format = '0.00'
            row += 1

        ws.freeze_panes = 'A2'
        wb.save(output_filepath)
        print(f"Excel file generated: {output_filepath}")

    # ------------------------ Main Processing ------------------------
    def remove_day1_early_punches(self, punches, cutoff=time(7, 15)):
        cutoff_mins = self.time_to_minutes(cutoff)
        return [p for p in punches if self.time_to_minutes(p) >= cutoff_mins]

    def process(self, day1_file, day2_file, output_file):
        print("Reading Day 1 attendance...")
        self.read_attendance_file(day1_file, 'day1')
        print("Reading Day 2 attendance...")
        self.read_attendance_file(day2_file, 'day2')

        print("Detecting cross-midnight shifts...")
        midnight_boundary = time(7, 15)

        for emp_code in self.employees.keys():
            emp = self.employees[emp_code]
            emp['day1_punches'] = self.remove_day1_early_punches(emp['day1_punches'])
            day1_punches, day2_punches, cross_info = self.detect_cross_midnight_shift(
                emp['day1_punches'], emp['day2_punches']
            )
            emp['day1_punches'] = day1_punches
            emp['day2_punches'] = day2_punches
            emp['cross_day_shift'] = cross_info

            emp['day1_hours'] = self.calculate_working_hours_total_span(day1_punches, 'day1')
            emp['day2_hours'] = self.calculate_working_hours_total_span(day2_punches, 'day2')

        print(f"Total employees processed: {len(self.employees)}")
        print("Generating Excel file...")
        self.generate_excel(output_file)
        print("Done!")

# ------------------------ Main Script ------------------------
def main():
    import sys
    processor = AttendanceProcessor()

    if len(sys.argv) > 1 and sys.argv[1] == '--analyze-punches':
        if len(sys.argv) < 3:
            print("Usage: python attendance_processor.py --analyze-punches '<punch_string>' [--breaks '<break_string>']")
            sys.exit(1)

        punch_str = sys.argv[2]
        custom_breaks = None
        if len(sys.argv) > 3 and sys.argv[3] == '--breaks' and len(sys.argv) > 4:
            breaks_str = sys.argv[4]
            if breaks_str.strip():
                custom_breaks = processor.parse_breaks(breaks_str)
            else:
                custom_breaks = []

        punches = processor.parse_punch_records(punch_str)
        processor.print_punch_analysis(punches, f"Custom Punches: {punch_str}", custom_breaks)
    else:
        processor.process(
            r'c:\contract\day1.txt',
            r'c:\contract\day2.txt',
            r'c:\contract\Desiredoutput_final_new.xlsx'
        )

if __name__ == '__main__':
    main()